# test case'ai:
# 1. svarus grazus failiukas be jokiu balaganu
# 2. failiukas su vienu pasikartojanciu codelistu
# 3. failiukas su keliais pasikartojanciais codelistais
# 4. failiukas su vienu pasikartojanciu codelistu ir clashais jame
# 5. failiukas su keliais pasikartojanciais codelistais ir clashais visuose juose

import os
import time
import xml.etree.ElementTree as et

from bs4 import BeautifulSoup as bs
from colorama import init, Back, Fore
from freg_funkcijos import normalize_text, print_xml, openxml, register_namespaces, remove_version_et, remove_version_str, sortCode
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from xml.etree.ElementTree import Element, ElementTree

def sublist(sublist, full_list): # patikrina ar sublist masyvas yra full_list masyvo poaibis
    if sublist == []:
        return False
    lst1 = list(set(full_list)) # kad jau tikrai sąrašas būtų
    lst2 = list(set(sublist))
    lst1 = [print_xml(a) for a in lst1]
    lst2 = [print_xml(a) for a in lst2]
    if all(a in lst1 for a in lst2):
        return True
    return False

def ets_equal(et1, et2):
    tag1 = et1.tag
    tag2 = et2.tag
    if tag1 != tag2:
        return False

    attrib1 = str(et1.attrib) # reikia sukurti kopiją, kitaip modifikuojamas originalus objektas
    attrib2 = str(et2.attrib)
    attrib1 = remove_version_str(attrib1) # nustatoma 1.0 versija
    attrib2 = remove_version_str(attrib2)
    if attrib1 != attrib2:
        return False

    text1 = normalize_text(et1.text)
    text2 = normalize_text(et2.text)
    if text1 != text2:
        return False

    if (list(et1) == []) ^ (list(et2) == []): # jei vienas turi child elementų, o kitas ne
        return False

    if (list(et1) != []) and (list(et2) != []): # jei abu turi child elementų
        if not children_equal(et1, et2): # jei tie child elementai nėra lygūs
            return False

    return True

def children_equal(et1, et2):

    children = list(et1) + list(et2) # bendras visų vaikų sąrašas
    total_len = len(children)

    for i, child in enumerate(children):
        for n in range(i+1, len(children)):
            if ets_equal(child, children[n]):
                children[i] = 0 # dublikatams nustatoma žinoma reikšmė

    while 0 in children:
        children.remove(0) # žinoma reikšmė pašalinama
        if 2 * len(children) == total_len: # ar et1 ir et2 vaikai sutapo
            return True
            
    return False

def conflict(et1, et2): # ar et1 ir et2 aprašo tą pačią rakto reikšmę (t.y. konfliktuoja vienas su kitu)
    try: # gali mesti exception, kai bandoma pasiekti neegzistuojantį atributą (šiuo atveju 'value')
        value1 = et1.attrib['value']
        value2 = et2.attrib['value']
        if value1 == value2:
            return True
    except:
        try:
            if et1.tag == et2.tag:
                lang1 = et1.attrib
                lang2 = et2.attrib
                if lang1 == lang2:
                    return True
        except:
            return False
        return False
    return False

def parse_xml_codelist(codelists, id):
    descriptions = []
    urns = []
    
    for codelist in codelists:   # codelist === vienos codelist versijos kodai (įrašai) -> "urn:sdmx:org.sdmx.infomodel.codelist.Codelist=LB:KS_APREPTIS_UVR(1.0).E"
        try:
            urns.append(codelist)
        except Exception as e:
            print('freg.py exception 1: ', e)
            pass

    urns.insert(0, descriptions) # kiekvienas description masyvas yra urns masyvo dalis (masyvų masyvas) ir yra vienos versijos codelistas

    descriptions = []
    conflicts = {}

    for codelist in urns: # skirtingos versijos su tuo pačiu id
        for code in codelist: # eina per visus vienos versijos įrašus
            flag = 0
            for element in descriptions: # tikrina, ar jau nėra tokio paties įrašo
                if ets_equal(element, code):
                    flag = 1

                elif conflict(element, code):
                    # patikrinti, ar vaikai lygūs
                    print(element.attrib, code.attrib)
                    children = list(element) + list(code)
                    print(children)

                    for i, child in enumerate(children):
                        for n in range(i+1, len(children)):
                            if ets_equal(child, children[n]):
                                children[i] = 0

                    while 0 in children:
                        children.remove(0)

                    if sublist(children, list(element)) or children == list(element):
                        flag = 1
                        break

                    elif sublist(children, list(code)) or children == list(code):
                        flag = 0
                        descriptions.remove(element)
                        break

                    # čia reikia padaryti, kad generuotų žodyną su konfliktais (KS_APREPTIS_UVR.A : [conflicting_element1, conflicting_element2, conflicting_element3])
                    conflict_id = code.attrib['urn']
                    conflict_id = conflict_id.split(':')[-1]
                    conflict_id = remove_version_str(conflict_id)

                    if conflict_id in conflicts:
                        if not any(ets_equal(code, elem) for elem in conflicts[conflict_id]): # ets_equal grąžina false, nors juos tiesiog reikia sumerginti
                            conflicts[conflict_id].append(code)
                    else:
                        conflicts[conflict_id] = [element, code]
                    flag = 1
            if flag == 0:
                descriptions.append(code) # tvarkingai surūšiuotas masyvas nuo didžiausios versijos iki mažiausios
    descriptions.sort(key = sortCode)
    # gal tiesiog geriau grąžinti vieną jau paruoštą codelistą ir jį appendinti prie codelists childo (ir removint visus kitus pradinius codelistus su tuo id)????

    return descriptions, conflicts

def main(filename):
    init()
    rt = openxml(filename)
    _, codelists = list(rt)
    
    versions = {}
    final_conflict_array = []

    if not os.path.isdir('excel_raktai'):
        os.mkdir('excel_raktai')

    for codelist in codelists: # kiekvienai rakto versijai sukuriamas atskiras codelistas

        # visus codelistus su tuo pačiu id sugrupuoja į vieną masyvą
        id = codelist.attrib['id']
        if id in versions:
            versions[id].append(codelist)
        else:
            versions[id] = [codelist]

    for id in versions:
        new_codelist, conflicts = parse_xml_codelist(versions[id], id)

        # atsikrato visų versijų, išskyrus pirmą sąraše
        for version in versions[id][1:]:
            codelists.remove(version)
        
        # nustato pirmos versijos atributus (kad likusi versija tikrai būtų pirmoji)
        versions[id][0].attrib['version'] = '1.0'
        versions[id][0].attrib['urn'] = remove_version_str(versions[id][0].attrib['urn']) # nustato versiją į 1.0
        versions[id][0].attrib['isFinal'] = 'true'
        
        # iš likusios versijos codelisto išmeta VISUS kodus
        for code in reversed(versions[id][0]):
            versions[id][0].remove(code)

        # prie jau tuščios versijos prideda naujus ir sutvarkytus kodus
        for code in new_codelist:
            versions[id][0].append(remove_version_et(code))

        # sukuriamas parent, child masyvas ir pridedamas prie final_conflict_array masyvo
        parent = versions[id][0]
        if len(conflicts) > 0:
            final_conflict_array.append((parent, conflicts))


    for parent, conflicts in final_conflict_array:

        wb = Workbook()
        wb.create_sheet('Pavadinimai')
        wb.create_sheet('XML')
        wb.remove(wb['Sheet'])

        ws2 = wb['Pavadinimai']
        ws_xml = wb['XML']
        
        id = parent.attrib['id']
        for conflict_id in conflicts:
            ws_xml.append([conflict_id])
            key_values = []
            key_lt = []
            key_en = []

            for elem in reversed(conflicts[conflict_id]): # elem = KS_APREPTIS_UVR(1.0).F. reversed tam, kad pirmiausia dėtų naujausią elementą
                text = print_xml(elem)
                elem_id = elem.get('urn').split('(')[-1]
                elem_id = elem_id.split(')')[0]

                for description in elem:
                    att = description.attrib
                    record = description.text

                    for lang in att:  # att dydis visada yra 1
                        language = att[lang]

                        if language.lower() == 'en':
                            key_en.append(elem_id)
                            key_en.append(record)
                            text = text.replace(record, '###en###')
                        elif language.lower() == 'lt':
                            key_lt.append(elem_id)
                            key_lt.append(record)
                            text = text.replace(record, '###lt###')

                if len(key_values) != 0:
                    if len(key_values[0]) < len(text):
                        key_values = []
                        key_values.append(text)
                else:
                    key_values.append(text)

            ws2.append([conflict_id.split('.')[-1]])
            ws2.append(key_lt)
            ws2.append(key_en)
            ws_xml.append(key_values)

        for i, row in enumerate(ws_xml.iter_rows()):
            for cell in row:
                cell.alignment = Alignment(wrap_text = True) # nustatoma, kad viename Excel langelyje automatiškai perkeltų tekstą į kitą eilutę, taip nusistato ir stulpelio aukštis

        for col in ws_xml.iter_cols():
            ws_xml.column_dimensions[col[0].column_letter].width = 110 # nustatomas langelio plotis

        for col in ws2.iter_cols():
            if col[0].column % 2 == 0:
                ws2.column_dimensions[col[0].column_letter].width = 70 # nustatomas langelio plotis

        for row in ws2.iter_rows():
            keys_lt = []
            keys_en = []
            ws2.protection.sheet = True
            ws2.protection.enable()
            

            if row[0].row % 3 == 1:
                ws2.merge_cells('%s:%s' % (row[0].coordinate, row[ws2.max_column - 1].coordinate))
                

            for cell in row:
                if cell.column % 2 == 0:
                    cell.protection = Protection(locked = False)

                if cell.row % 3 == 1:
                    cell.fill = PatternFill(fgColor = 'c2ffd1', fill_type = 'solid')
                    cell.font = Font(size = 25, bold = True)
                    ws2.row_dimensions[cell.row].height = 30
                    cell.alignment = Alignment(horizontal = 'left')

                elif cell.row % 3 == 2:
                    if cell.value not in keys_lt and cell.column % 2 == 0:
                        if keys_lt != []:
                            cell.font = Font(color = 'ff0000')
                        keys_lt.append(cell.value)

                else:
                    if cell.value not in keys_en and cell.column % 2 == 0:
                        if keys_en != []:
                            cell.font = Font(color = 'ff0000')
                        keys_en.append(cell.value)

        ws_xml.protection.sheet = True
        ws_xml.protection.enable()
        wb.save(os.path.join('excel_raktai', ('%s.xlsx' % id)))    

    final_string = et.tostring(rt)
    with open('new_' + filename, 'wb') as f:
        f.write(final_string) # išsaugomas xml failas, tačiau vietoje konfliktų prisegama pirma versija (vėliau parenkama, kokia norima)

    print(Back.GREEN + 'Programa baigė darbą sėkmingai' + Back.BLACK)

main('small_codelist.xml')
