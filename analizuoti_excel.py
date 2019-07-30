import os
import xml.etree.ElementTree as et
import time

from bs4 import BeautifulSoup as bs
from colorama import init, Back, Fore
from freg_funkcijos import normalize_text, openxml, print_xml, register_namespaces, remove_version_et, remove_version_str, sortCode
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from xml.etree.ElementTree import Element, ElementTree

def parse_str_to_element(child):
    if '######' not in child:
        elem = et.fromstring(child)
        elem = remove_version_et(elem)
        return [elem]
    else:
        children = child.split('######')
        to_return = []
        for ch in children:
            elem = et.fromstring(ch)
            elem = remove_version_et(elem)
            to_return.append(elem)

        return to_return
        


def add_child_to_codelist(codelists, parent_id, child):
    for codelist in codelists:
        for code in codelist:
            try:
                if parent_id in code.get('urn'):
                    codelist.remove(code)
                    codelist.append(child)
                    break
            except:
                for key in code.attrib:
                    if code.attrib[key] in (child.attrib).values():
                        if parent_id.split('.')[0] in codelist.attrib['id']:
                            codelist.remove(code)
                            codelist.append(child)
                            break

def main(file):
    init()
    rt = openxml('new_' + file)
    _, codelists = list(rt)

    filenames = os.listdir('excel_raktai')

    for filename in filenames:
        wb = load_workbook(os.path.join('excel_raktai', filename))
        ws_xml = wb['XML']
        ws_pav = wb['Pavadinimai']

        rows = list(ws_pav.iter_rows())
        keys = {}
        for i, row in enumerate(rows):
            key_lt = ''
            key_en = ''

            if i % 3 == 0: # parent elemento ID
                parent_id = row[0].value
                if 'pavadinimas' not in parent_id:
                    parent_id = filename.split('.')[0] + '(1.0).' + parent_id
                else:
                    parent_id = filename.split('.')[0] + '.' + parent_id
                keys[parent_id] = {}

            elif i % 3 == 1: # lietuviška rakto reikšmė
                for cell in reversed(row):
                    if cell.value != None and cell.column % 2 == 0:
                        keys[parent_id]['lt'] = cell.value
            else: # angliška rakto reikšmė
                for cell in reversed(row):
                    if cell.value != None and cell.column % 2 == 0:
                        keys[parent_id]['en'] = cell.value

        rows = list(ws_xml.iter_rows())
        for i, row in enumerate(rows):
            if i % 2 == 0:
                parent_id = row[0].value
            else:
                child = str(row[0].value)

                if 'lt' in keys[parent_id] and 'en' in keys[parent_id]:
                    child = child.replace('###lt###', normalize_text(keys[parent_id]['lt']))
                    child = child.replace('###en###', normalize_text(keys[parent_id]['en']))

                elif 'lt' in keys[parent_id] and not 'en' in keys[parent_id]: # jei egzistuoja tik lietuviška rakto reikšmės versija
                    child = child.replace('###lt###', normalize_text(keys[parent_id]['lt']))
                    print('Galima klaida „%s“. Nėra angliško kodo aprašymo' % parent_id)

                elif 'en' in keys[parent_id] and not 'lt' in keys[parent_id]: # jei egzistuoja tik angliška rakto reikšmės versija
                    child = child.replace('###en###', normalize_text(keys[parent_id]['en']))
                    print('Galima klaida „%s“. Nėra lietuviško kodo aprašymo' % parent_id)

                else:
                    print('Klaida „%s“. Nėra reikšmės nei viena kalba' % parent_id)
                    
                # čia pakeisti xml failą
                children = parse_str_to_element(child)
                for child in children:
                    add_child_to_codelist(codelists, parent_id, child)

    # išrikiuoti abėcėlės tvarka

    for codelist in codelists:
        codes = []
        for code in codelist:
            codes.append(code) # sukuria sąrašą su visais jau esamais kodais

        codes.sort(key = sortCode) 

        for code in codes:
            codelist.append(code) # prie seno kodų sąrašo prideda naują, surikiuotą sąrašą

        total_codes = len(codes)
        
        for _ in range(total_codes):
            codelist.remove(codelist[0]) # lauk išmetami nesurikiuoti kodai (kai nulinis nuimamas tuomet viskas pasislenka per vieną)

    # ir jį išsaugoti
    final_string = et.tostring(rt)
    with open('new_final_' + file, 'wb') as f:
        f.write(final_string)

    print(Back.GREEN + 'Failas išsaugotas sėkmingai' + Back.BLACK)
    #_ = input()

main('codelist.xml')
